"""
ETL: CSV crudo C5 CDMX → esquema estrella (fact + 5 dims + diccionario).

Uso:
    python -m etl.build_star_schema \
        --input  data/c5_raw/inViales_2022_2024.csv \
        --output data/c5_processed \
        [--sample 50000]
"""

from __future__ import annotations

import argparse
import math
import os
import sys
import unicodedata
import warnings
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import requests

warnings.filterwarnings("ignore", category=FutureWarning)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────
CDMX_LAT_MIN, CDMX_LAT_MAX = 19.04, 19.59
CDMX_LON_MIN, CDMX_LON_MAX = -99.36, -98.94

OPEN_METEO_URL = "https://archive-api.open-meteo.com/v1/archive"
OPEN_METEO_VARS = (
    "temperature_2m,precipitation,visibility,"
    "relative_humidity_2m,weathercode"
)
CDMX_LAT_CENTER, CDMX_LON_CENTER = 19.4326, -99.1332

SENTINEL_SK = -1


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _strip_accents(text: str) -> str:
    """Elimina tildes y normaliza a ASCII."""
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def _normalize_alcaldia(series: pd.Series) -> pd.Series:
    return (
        series.fillna("DESCONOCIDA")
        .astype(str)
        .str.strip()
        .str.upper()
        .apply(_strip_accents)
    )


def _read_raw_csv(path: Path, sample: Optional[int]) -> pd.DataFrame:
    """Lee el CSV con fallback de encoding; opcionalmente muestra una muestra."""
    for enc in ("utf-8", "latin-1"):
        try:
            df = pd.read_csv(
                path,
                encoding=enc,
                low_memory=False,
                nrows=sample,
            )
            print(f"  CSV leído con encoding={enc}  ({len(df):,} filas)")
            return df
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"No se pudo leer {path} con utf-8 ni latin-1.")


def _parse_datetime(date_col: pd.Series, time_col: pd.Series) -> pd.Series:
    combined = date_col.astype(str).str.strip() + " " + time_col.astype(str).str.strip()
    return pd.to_datetime(combined, errors="coerce")


# ─────────────────────────────────────────────────────────────────────────────
# 1. PREPROCESAMIENTO
# ─────────────────────────────────────────────────────────────────────────────

def preprocess(df: pd.DataFrame) -> pd.DataFrame:
    n0 = len(df)

    df["datetime_creacion"] = _parse_datetime(df["fecha_creacion"], df["hora_creacion"])
    df["datetime_cierre"] = _parse_datetime(df["fecha_cierre"], df["hora_cierre"])

    # Filtrar nulos en datetime_creacion
    df = df[df["datetime_creacion"].notna()].copy()
    n1 = len(df)

    # Filtrar coordenadas fuera de CDMX
    lat = pd.to_numeric(df["latitud"], errors="coerce")
    lon = pd.to_numeric(df["longitud"], errors="coerce")
    mask = (
        lat.between(CDMX_LAT_MIN, CDMX_LAT_MAX) &
        lon.between(CDMX_LON_MIN, CDMX_LON_MAX)
    )
    df = df[mask].copy()
    df["latitud"] = lat[mask]
    df["longitud"] = lon[mask]
    n2 = len(df)

    # Normalizar alcaldías
    df["alcaldia_inicio"] = _normalize_alcaldia(df["alcaldia_inicio"])
    df["alcaldia_cierre"] = _normalize_alcaldia(df["alcaldia_cierre"])
    df["alcaldia_catalogo"] = _normalize_alcaldia(df["alcaldia_catalogo"])

    print(f"  Preprocesamiento: {n0:,} → {n1:,} (drop datetime nulo) → {n2:,} (drop coords CDMX)")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 2. DIM_TIEMPO
# ─────────────────────────────────────────────────────────────────────────────

def build_dim_tiempo(df: pd.DataFrame) -> pd.DataFrame:
    horas_creacion = df["datetime_creacion"].dt.floor("h")
    horas_cierre = df["datetime_cierre"].dropna().dt.floor("h")
    all_hours = pd.concat([horas_creacion, horas_cierre]).drop_duplicates().dropna().sort_values()

    dt = pd.DataFrame({"fecha_hora": all_hours.values})
    dt = dt.drop_duplicates(subset="fecha_hora").reset_index(drop=True)
    dt["sk_tiempo"] = dt.index + 1

    s = dt["fecha_hora"]
    dt["año"] = s.dt.year
    dt["mes"] = s.dt.month
    dt["dia"] = s.dt.day
    dt["hora"] = s.dt.hour
    dt["dia_semana"] = s.dt.dayofweek          # 0=lunes, 6=domingo
    dt["nombre_dia"] = s.dt.day_name(locale=None)
    dt["trimestre"] = s.dt.quarter
    dt["semana_del_año"] = s.dt.isocalendar().week.astype(int)
    dt["es_finde"] = dt["dia_semana"] >= 5
    dt["es_noche"] = (dt["hora"] >= 20) | (dt["hora"] < 6)
    dt["es_hora_pico"] = (
        (~dt["es_finde"]) &
        (
            dt["hora"].between(7, 9) |
            dt["hora"].between(17, 19)
        )
    )

    def _turno(h):
        if 0 <= h < 6:
            return "madrugada"
        if 6 <= h < 12:
            return "manana"
        if 12 <= h < 20:
            return "tarde"
        return "noche"

    dt["turno"] = dt["hora"].map(_turno)

    dt["hora_sin"] = np.sin(2 * math.pi * dt["hora"] / 24)
    dt["hora_cos"] = np.cos(2 * math.pi * dt["hora"] / 24)
    dt["dia_sin"] = np.sin(2 * math.pi * dt["dia_semana"] / 7)
    dt["dia_cos"] = np.cos(2 * math.pi * dt["dia_semana"] / 7)

    cols = [
        "sk_tiempo", "fecha_hora", "año", "mes", "dia", "hora",
        "dia_semana", "nombre_dia", "trimestre", "semana_del_año",
        "es_finde", "es_noche", "es_hora_pico", "turno",
        "hora_sin", "hora_cos", "dia_sin", "dia_cos",
    ]
    return dt[cols]


# ─────────────────────────────────────────────────────────────────────────────
# 3. DIM_UBICACION
# ─────────────────────────────────────────────────────────────────────────────

def build_dim_ubicacion(df: pd.DataFrame) -> pd.DataFrame:
    tmp = df.copy()
    tmp["lat_idx"] = (tmp["latitud"] * 100).astype(int)
    tmp["lon_idx"] = (tmp["longitud"] * 100).astype(int)
    tmp["cuadrante_id"] = "CDMX_" + tmp["lat_idx"].astype(str) + "_" + tmp["lon_idx"].astype(str)

    tmp["alcaldia"] = tmp["alcaldia_catalogo"].where(
        tmp["alcaldia_catalogo"] != "DESCONOCIDA", tmp["alcaldia_inicio"]
    )
    tmp["colonia"] = tmp["colonia_catalogo"].fillna("DESCONOCIDA").astype(str).str.strip()

    grp = (
        tmp.groupby(["cuadrante_id", "alcaldia", "colonia"])
        .agg(lat_centro=("latitud", "mean"), lon_centro=("longitud", "mean"))
        .reset_index()
    )
    grp["sk_ubicacion"] = grp.index + 1

    cols = ["sk_ubicacion", "cuadrante_id", "lat_centro", "lon_centro", "alcaldia", "colonia"]
    return grp[cols]


# ─────────────────────────────────────────────────────────────────────────────
# 4. DIM_CLIMA  (Open-Meteo Archive, con cache y fallback gracioso)
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_meteo_year(year: int, cache_dir: Path) -> Optional[pd.DataFrame]:
    cache_file = cache_dir / f"clima_{year}.csv"
    if cache_file.exists():
        print(f"    [clima] cache hit: {cache_file.name}")
        return pd.read_csv(cache_file, parse_dates=["fecha_hora"])

    start = f"{year}-01-01"
    end = f"{year}-12-31"
    params = {
        "latitude": CDMX_LAT_CENTER,
        "longitude": CDMX_LON_CENTER,
        "start_date": start,
        "end_date": end,
        "hourly": OPEN_METEO_VARS,
        "timezone": "America/Mexico_City",
    }
    try:
        resp = requests.get(OPEN_METEO_URL, params=params, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        hourly = data["hourly"]
        chunk = pd.DataFrame({
            "fecha_hora": pd.to_datetime(hourly["time"]),
            "temperatura_c": hourly["temperature_2m"],
            "precipitacion_mm": hourly["precipitation"],
            "visibilidad_km": pd.Series(hourly["visibility"]).div(1000),
            "humedad_pct": hourly["relative_humidity_2m"],
            "codigo_clima": hourly["weathercode"],
        })
        chunk.to_csv(cache_file, index=False)
        print(f"    [clima] descargado y cacheado: {cache_file.name}  ({len(chunk):,} horas)")
        return chunk
    except Exception as exc:
        print(f"\n  ⚠️  WARNING [clima] año {year}: {exc}")
        return None


def build_dim_clima(df: pd.DataFrame, cache_dir: Path) -> tuple[pd.DataFrame, bool]:
    cache_dir.mkdir(parents=True, exist_ok=True)

    min_year = df["datetime_creacion"].dt.year.min()
    max_year = df["datetime_creacion"].dt.year.max()

    chunks = []
    api_ok = True
    for yr in range(int(min_year), int(max_year) + 1):
        chunk = _fetch_meteo_year(yr, cache_dir)
        if chunk is None:
            api_ok = False
        else:
            chunks.append(chunk)

    # Construir todas las horas del dataset para tener la dimensión completa
    horas_creacion = df["datetime_creacion"].dt.floor("h")
    horas_cierre = df["datetime_cierre"].dropna().dt.floor("h")
    all_hours = (
        pd.concat([horas_creacion, horas_cierre])
        .drop_duplicates()
        .dropna()
        .sort_values()
        .reset_index(drop=True)
    )
    base = pd.DataFrame({"fecha_hora": all_hours.values})

    if chunks:
        clima_raw = pd.concat(chunks, ignore_index=True)
        clima_raw["fecha_hora"] = clima_raw["fecha_hora"].dt.tz_localize(None)
        merged = base.merge(clima_raw, on="fecha_hora", how="left")
    else:
        merged = base.copy()
        for col in ["temperatura_c", "precipitacion_mm", "visibilidad_km", "humedad_pct", "codigo_clima"]:
            merged[col] = np.nan

    def _cat_temp(t):
        if pd.isna(t):
            return None
        return "frio" if t < 15 else ("templado" if t <= 25 else "calido")

    merged["categoria_temp"] = merged["temperatura_c"].apply(_cat_temp)
    merged["hay_lluvia"] = merged["precipitacion_mm"].fillna(0) > 0
    merged["visibilidad_baja"] = merged["visibilidad_km"].fillna(np.inf) < 1

    merged = merged.drop_duplicates(subset="fecha_hora").reset_index(drop=True)
    merged["sk_clima"] = merged.index + 1

    cols = [
        "sk_clima", "fecha_hora", "temperatura_c", "precipitacion_mm",
        "visibilidad_km", "humedad_pct", "codigo_clima",
        "categoria_temp", "hay_lluvia", "visibilidad_baja",
    ]
    return merged[cols], api_ok


# ─────────────────────────────────────────────────────────────────────────────
# 5. DIM_CALENDARIO
# ─────────────────────────────────────────────────────────────────────────────

def build_dim_calendario(df: pd.DataFrame) -> pd.DataFrame:
    import holidays as hol_lib

    min_date = df["datetime_creacion"].dt.date.min()
    max_date = df["datetime_creacion"].dt.date.max()
    all_dates = pd.date_range(start=min_date, end=max_date, freq="D")

    min_year = min_date.year
    max_year = max_date.year
    mx_holidays: dict = {}
    for yr in range(min_year, max_year + 1):
        mx_holidays.update(hol_lib.Mexico(years=yr))

    cal = pd.DataFrame({"fecha": all_dates.date})
    cal["fecha_dt"] = pd.to_datetime(cal["fecha"])
    cal["dow"] = cal["fecha_dt"].dt.dayofweek

    cal["es_festivo"] = cal["fecha"].apply(lambda d: d in mx_holidays)
    cal["nombre_festivo"] = cal["fecha"].apply(lambda d: mx_holidays.get(d, None))

    def _dia_tipo(row):
        if row["es_festivo"]:
            return "festivo"
        if row["dow"] >= 5:
            return "finde"
        return "laboral"

    cal["dia_tipo"] = cal.apply(_dia_tipo, axis=1)

    cal["es_quincena"] = cal["fecha_dt"].apply(
        lambda d: d.day == 15 or d.day == (d + pd.offsets.MonthEnd(0)).day
    )

    def _es_vacaciones(d: pd.Timestamp) -> bool:
        m, day = d.month, d.day
        # Julio completo
        if m == 7:
            return True
        # Primera quincena agosto
        if m == 8 and day <= 15:
            return True
        # Segunda quincena diciembre
        if m == 12 and day >= 16:
            return True
        # Primera quincena enero
        if m == 1 and day <= 15:
            return True
        # Semana santa: última semana de marzo o primera de abril (aprox)
        if m == 3 and day >= 24:
            return True
        if m == 4 and day <= 7:
            return True
        return False

    cal["es_vacaciones"] = cal["fecha_dt"].apply(_es_vacaciones)
    cal = cal.drop(columns=["fecha_dt", "dow"]).reset_index(drop=True)
    cal["sk_calendario"] = cal.index + 1

    cols = ["sk_calendario", "fecha", "dia_tipo", "es_festivo", "nombre_festivo", "es_quincena", "es_vacaciones"]
    return cal[cols]


# ─────────────────────────────────────────────────────────────────────────────
# 6. DIM_TIPO_INCIDENTE
# ─────────────────────────────────────────────────────────────────────────────

def build_dim_tipo_incidente(df: pd.DataFrame) -> pd.DataFrame:
    cols_src = ["tipo_incidente_c4", "incidente_c4", "tipo_entrada"]
    tmp = df[cols_src].copy()
    for c in cols_src:
        tmp[c] = tmp[c].fillna("DESCONOCIDO").astype(str).str.strip()

    dim = tmp.drop_duplicates().reset_index(drop=True)
    dim["sk_tipo"] = dim.index + 1
    return dim[["sk_tipo"] + cols_src]


# ─────────────────────────────────────────────────────────────────────────────
# 7. FACT_INCIDENTES
# ─────────────────────────────────────────────────────────────────────────────

def build_fact(
    df: pd.DataFrame,
    dim_tiempo: pd.DataFrame,
    dim_ubicacion: pd.DataFrame,
    dim_clima: pd.DataFrame,
    dim_calendario: pd.DataFrame,
    dim_tipo: pd.DataFrame,
) -> pd.DataFrame:

    fact = df.copy()
    fact["sk_incidente"] = np.arange(1, len(fact) + 1)

    # ── tiempo ────────────────────────────────────────────────────────────────
    tiempo_lookup = dim_tiempo.set_index("fecha_hora")["sk_tiempo"]

    fact["_hora_creacion"] = fact["datetime_creacion"].dt.floor("h")
    fact["_hora_cierre"] = fact["datetime_cierre"].dt.floor("h")

    fact["fk_tiempo_creacion"] = fact["_hora_creacion"].map(tiempo_lookup)
    fact["fk_tiempo_cierre"] = fact["_hora_cierre"].map(tiempo_lookup)

    # ── ubicacion ─────────────────────────────────────────────────────────────
    fact["lat_idx"] = (fact["latitud"] * 100).astype(int)
    fact["lon_idx"] = (fact["longitud"] * 100).astype(int)
    fact["cuadrante_id"] = "CDMX_" + fact["lat_idx"].astype(str) + "_" + fact["lon_idx"].astype(str)

    fact["alcaldia"] = fact["alcaldia_catalogo"].where(
        fact["alcaldia_catalogo"] != "DESCONOCIDA", fact["alcaldia_inicio"]
    )
    fact["colonia"] = fact["colonia_catalogo"].fillna("DESCONOCIDA").astype(str).str.strip()

    ubic_lookup = dim_ubicacion.set_index(
        ["cuadrante_id", "alcaldia", "colonia"]
    )["sk_ubicacion"]
    fact["fk_ubicacion"] = pd.MultiIndex.from_arrays(
        [fact["cuadrante_id"], fact["alcaldia"], fact["colonia"]]
    ).map(ubic_lookup)

    # ── clima ──────────────────────────────────────────────────────────────────
    clima_lookup = dim_clima.set_index("fecha_hora")["sk_clima"]
    fact["fk_clima"] = fact["_hora_creacion"].map(clima_lookup)

    # ── calendario ─────────────────────────────────────────────────────────────
    cal_lookup = dim_calendario.set_index("fecha")["sk_calendario"]
    fact["_fecha_creacion"] = fact["datetime_creacion"].dt.date
    fact["fk_calendario"] = fact["_fecha_creacion"].map(cal_lookup)

    # ── tipo incidente ─────────────────────────────────────────────────────────
    for c in ["tipo_incidente_c4", "incidente_c4", "tipo_entrada"]:
        fact[c] = fact[c].fillna("DESCONOCIDO").astype(str).str.strip()
    tipo_lookup = dim_tipo.set_index(
        ["tipo_incidente_c4", "incidente_c4", "tipo_entrada"]
    )["sk_tipo"]
    fact["fk_tipo_incidente"] = pd.MultiIndex.from_arrays(
        [fact["tipo_incidente_c4"], fact["incidente_c4"], fact["tipo_entrada"]]
    ).map(tipo_lookup)

    # ── métricas derivadas ─────────────────────────────────────────────────────
    diff_min = (fact["datetime_cierre"] - fact["datetime_creacion"]).dt.total_seconds() / 60
    fact["tiempo_respuesta_min"] = diff_min.where(diff_min.between(0, 1440), other=np.nan)
    fact["tiempo_respuesta_min"] = fact["tiempo_respuesta_min"].round(0).astype("Int64")

    fact["codigo_cierre"] = (
        fact["codigo_cierre"].fillna("desconocido").astype(str).str.strip().str.lower()
    )

    fact["es_falsa_alarma"] = (
        fact["clas_con_f_alarma"].astype(str).str.upper().str.contains("FALSA", na=False)
    )

    fact["mismo_lugar_inicio_cierre"] = (
        fact["alcaldia_inicio"] == fact["alcaldia_cierre"]
    )

    cols_out = [
        "sk_incidente", "folio",
        "fk_tiempo_creacion", "fk_tiempo_cierre",
        "fk_ubicacion", "fk_clima", "fk_calendario", "fk_tipo_incidente",
        "latitud", "longitud",
        "tiempo_respuesta_min", "codigo_cierre",
        "es_falsa_alarma", "mismo_lugar_inicio_cierre",
    ]
    return fact[cols_out]


# ─────────────────────────────────────────────────────────────────────────────
# 8. INTEGRIDAD REFERENCIAL CON SENTINEL -1
# ─────────────────────────────────────────────────────────────────────────────

FK_MAP = {
    "fk_tiempo_creacion": ("dim_tiempo", "sk_tiempo"),
    "fk_tiempo_cierre":   ("dim_tiempo", "sk_tiempo"),
    "fk_ubicacion":       ("dim_ubicacion", "sk_ubicacion"),
    "fk_clima":           ("dim_clima", "sk_clima"),
    "fk_calendario":      ("dim_calendario", "sk_calendario"),
    "fk_tipo_incidente":  ("dim_tipo_incidente", "sk_tipo"),
}


def fix_referential_integrity(
    fact: pd.DataFrame,
    dims: dict[str, pd.DataFrame],
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    dims = {k: v.copy() for k, v in dims.items()}
    orphan_report = {}

    for fk_col, (dim_name, sk_col) in FK_MAP.items():
        dim = dims[dim_name]
        valid_sks = set(dim[sk_col].dropna())

        mask_null = fact[fk_col].isna()
        mask_orphan = (~mask_null) & (~fact[fk_col].isin(valid_sks))
        total_bad = mask_null.sum() + mask_orphan.sum()

        if total_bad > 0:
            orphan_report[fk_col] = total_bad
            fact.loc[mask_null | mask_orphan, fk_col] = SENTINEL_SK

            # Añadir registro sentinela si no existe ya
            if SENTINEL_SK not in valid_sks:
                sentinel_row = {col: (SENTINEL_SK if col == sk_col else "DESCONOCIDO")
                                for col in dim.columns}
                # Para columnas numéricas usar NaN
                for col in dim.columns:
                    if col == sk_col:
                        continue
                    if pd.api.types.is_numeric_dtype(dim[col]):
                        sentinel_row[col] = np.nan
                    elif pd.api.types.is_bool_dtype(dim[col]):
                        sentinel_row[col] = False
                dims[dim_name] = pd.concat(
                    [dim, pd.DataFrame([sentinel_row])], ignore_index=True
                )

    if orphan_report:
        print("\n  ⚠️  Integridad referencial — FKs con valores sentinela asignados:")
        for col, n in orphan_report.items():
            print(f"     {col}: {n:,} filas → sk={SENTINEL_SK}")
    else:
        print("  ✓  Integridad referencial OK (sin huérfanas)")

    # Convertir FKs a Int64 (nullable)
    for fk_col in FK_MAP:
        fact[fk_col] = fact[fk_col].astype("Int64")

    return fact, dims


# ─────────────────────────────────────────────────────────────────────────────
# 9. DICCIONARIO DE DATOS
# ─────────────────────────────────────────────────────────────────────────────

DICT_META: dict[str, list[dict]] = {
    "fact_incidentes": [
        ("sk_incidente",             "INTEGER", "Surrogate key autoincremental del hecho", "",                           True,  False, ""),
        ("folio",                    "VARCHAR", "Clave natural del incidente según C5 CDMX", "",                         False, False, ""),
        ("fk_tiempo_creacion",       "INTEGER", "FK al momento (hora) de creación del incidente", "",                    False, True,  "dim_tiempo.sk_tiempo"),
        ("fk_tiempo_cierre",         "INTEGER", "FK al momento (hora) de cierre del incidente", "",                      False, True,  "dim_tiempo.sk_tiempo"),
        ("fk_ubicacion",             "INTEGER", "FK al cuadrante geográfico de ~1km² donde ocurrió", "",                 False, True,  "dim_ubicacion.sk_ubicacion"),
        ("fk_clima",                 "INTEGER", "FK a las condiciones climáticas en la hora de creación", "",            False, True,  "dim_clima.sk_clima"),
        ("fk_calendario",            "INTEGER", "FK al día calendario de creación", "",                                  False, True,  "dim_calendario.sk_calendario"),
        ("fk_tipo_incidente",        "INTEGER", "FK al tipo/subtipo/canal de ingreso del incidente", "",                 False, True,  "dim_tipo_incidente.sk_tipo"),
        ("latitud",                  "FLOAT",   "Latitud WGS84 del incidente (para mapas)", "19.04–19.59",               False, False, ""),
        ("longitud",                 "FLOAT",   "Longitud WGS84 del incidente (para mapas)", "-99.36–-98.94",            False, False, ""),
        ("tiempo_respuesta_min",     "INTEGER", "Minutos entre creación y cierre; NULL si negativo o >1440", "0–1440",   False, False, ""),
        ("codigo_cierre",            "VARCHAR", "Resultado del cierre: afirmativo, negativo, falsa_alarma, etc.", "",    False, False, ""),
        ("es_falsa_alarma",          "BOOLEAN", "True si clas_con_f_alarma contiene 'FALSA'", "True/False",             False, False, ""),
        ("mismo_lugar_inicio_cierre","BOOLEAN", "True si alcaldía_inicio == alcaldía_cierre", "True/False",             False, False, ""),
    ],
    "dim_tiempo": [
        ("sk_tiempo",    "INTEGER",  "Surrogate key de la hora", "", True, False, ""),
        ("fecha_hora",   "DATETIME", "Timestamp truncado a la hora exacta (grano)", "", False, False, ""),
        ("año",          "INTEGER",  "Año calendario", "2021–2025", False, False, ""),
        ("mes",          "INTEGER",  "Mes del año (1–12)", "1–12", False, False, ""),
        ("dia",          "INTEGER",  "Día del mes", "1–31", False, False, ""),
        ("hora",         "INTEGER",  "Hora del día (0–23)", "0–23", False, False, ""),
        ("dia_semana",   "INTEGER",  "0=lunes, 6=domingo", "0–6", False, False, ""),
        ("nombre_dia",   "VARCHAR",  "Nombre del día en inglés (pandas default)", "", False, False, ""),
        ("trimestre",    "INTEGER",  "Trimestre del año", "1–4", False, False, ""),
        ("semana_del_año","INTEGER", "Semana ISO del año", "1–53", False, False, ""),
        ("es_finde",     "BOOLEAN",  "True si sábado o domingo", "True/False", False, False, ""),
        ("es_noche",     "BOOLEAN",  "True si hora entre 20:00 y 05:59", "True/False", False, False, ""),
        ("es_hora_pico", "BOOLEAN",  "True si lun–vie entre 7–9h o 17–19h", "True/False", False, False, ""),
        ("turno",        "VARCHAR",  "Turno del día según hora", "madrugada/manana/tarde/noche", False, False, ""),
        ("hora_sin",     "FLOAT",    "Encoding cíclico seno de la hora (sin(2π·h/24))", "-1–1", False, False, ""),
        ("hora_cos",     "FLOAT",    "Encoding cíclico coseno de la hora", "-1–1", False, False, ""),
        ("dia_sin",      "FLOAT",    "Encoding cíclico seno del día semana", "-1–1", False, False, ""),
        ("dia_cos",      "FLOAT",    "Encoding cíclico coseno del día semana", "-1–1", False, False, ""),
    ],
    "dim_ubicacion": [
        ("sk_ubicacion", "INTEGER", "Surrogate key del cuadrante-alcaldía-colonia", "", True, False, ""),
        ("cuadrante_id", "VARCHAR", "ID del cuadrante ~1km²: CDMX_{lat*100}_{lon*100}", "", False, False, ""),
        ("lat_centro",   "FLOAT",   "Latitud media de incidentes en el cuadrante", "", False, False, ""),
        ("lon_centro",   "FLOAT",   "Longitud media de incidentes en el cuadrante", "", False, False, ""),
        ("alcaldia",     "VARCHAR", "Alcaldía según catálogo; fallback a alcaldia_inicio", "", False, False, ""),
        ("colonia",      "VARCHAR", "Colonia según catálogo C5", "", False, False, ""),
    ],
    "dim_clima": [
        ("sk_clima",         "INTEGER", "Surrogate key de la observación climática horaria", "", True, False, ""),
        ("fecha_hora",       "DATETIME","Timestamp UTC-6 de la observación horaria (Open-Meteo)", "", False, False, ""),
        ("temperatura_c",    "FLOAT",   "Temperatura a 2m de altura en °C", "", False, False, ""),
        ("precipitacion_mm", "FLOAT",   "Precipitación acumulada en la hora en mm", "≥0", False, False, ""),
        ("visibilidad_km",   "FLOAT",   "Visibilidad horizontal en km", "≥0", False, False, ""),
        ("humedad_pct",      "FLOAT",   "Humedad relativa a 2m en %", "0–100", False, False, ""),
        ("codigo_clima",     "INTEGER", "WMO Weather Interpretation Code", "0–99", False, False, ""),
        ("categoria_temp",   "VARCHAR", "Categoría de temperatura", "frio/templado/calido", False, False, ""),
        ("hay_lluvia",       "BOOLEAN", "True si precipitacion_mm > 0", "True/False", False, False, ""),
        ("visibilidad_baja", "BOOLEAN", "True si visibilidad_km < 1", "True/False", False, False, ""),
    ],
    "dim_calendario": [
        ("sk_calendario",  "INTEGER", "Surrogate key del día calendario", "", True, False, ""),
        ("fecha",          "DATE",    "Fecha única por registro", "", False, False, ""),
        ("dia_tipo",       "VARCHAR", "Tipo de día", "laboral/finde/festivo", False, False, ""),
        ("es_festivo",     "BOOLEAN", "True si es festivo oficial México", "True/False", False, False, ""),
        ("nombre_festivo", "VARCHAR", "Nombre del festivo oficial o NULL", "", False, False, ""),
        ("es_quincena",    "BOOLEAN", "True si es día 15 o último del mes", "True/False", False, False, ""),
        ("es_vacaciones",  "BOOLEAN", "True si cae en periodo vacacional aproximado", "True/False", False, False, ""),
    ],
    "dim_tipo_incidente": [
        ("sk_tipo",           "INTEGER", "Surrogate key del tipo de incidente", "", True, False, ""),
        ("tipo_incidente_c4", "VARCHAR", "Categoría principal del incidente según clasificación C4", "", False, False, ""),
        ("incidente_c4",      "VARCHAR", "Descripción específica del incidente según C4", "", False, False, ""),
        ("tipo_entrada",      "VARCHAR", "Canal por el que se reportó el incidente", "call_center/app/boton/etc", False, False, ""),
    ],
}

TABLE_SOURCES = {
    "fact_incidentes":   "CSV crudo C5 CDMX + derivados ETL",
    "dim_tiempo":        "Derivado de datetime_creacion / datetime_cierre del CSV crudo",
    "dim_ubicacion":     "Derivado de latitud/longitud + alcaldia/colonia del CSV crudo",
    "dim_clima":         "Open-Meteo Archive API (lat=19.4326, lon=-99.1332)",
    "dim_calendario":    "Derivado de fechas del CSV crudo + librería holidays.Mexico()",
    "dim_tipo_incidente":"Derivado de tipo_incidente_c4 / incidente_c4 / tipo_entrada del CSV crudo",
}


def build_diccionario(
    out_path: Path,
    dims: dict[str, pd.DataFrame],
    fact: pd.DataFrame,
) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(color="FFFFFF", bold=True)

    tables_data = {
        "fact_incidentes": fact,
        "dim_tiempo": dims["dim_tiempo"],
        "dim_ubicacion": dims["dim_ubicacion"],
        "dim_clima": dims["dim_clima"],
        "dim_calendario": dims["dim_calendario"],
        "dim_tipo_incidente": dims["dim_tipo_incidente"],
    }

    dict_cols = ["tabla", "columna", "tipo_dato", "descripcion",
                 "valores_validos", "es_pk", "es_fk", "fk_referencia", "fuente"]

    for tbl_name, meta_rows in DICT_META.items():
        ws = wb.create_sheet(title=tbl_name[:31])
        ws.append(dict_cols)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        source = TABLE_SOURCES.get(tbl_name, "")
        for row_meta in meta_rows:
            col_name, dtype, desc, vals, is_pk, is_fk, fk_ref = row_meta
            ws.append([tbl_name, col_name, dtype, desc, vals, is_pk, is_fk, fk_ref, source])

        for col_idx, _ in enumerate(dict_cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Hoja resumen
    ws_res = wb.create_sheet(title="_RESUMEN")
    ws_res.append(["tabla", "n_filas", "n_columnas", "tamanio_mb"])
    for cell in ws_res[1]:
        cell.fill = header_fill
        cell.font = header_font

    for tbl_name, tbl_df in tables_data.items():
        size_mb = tbl_df.memory_usage(deep=True).sum() / (1024 ** 2)
        ws_res.append([tbl_name, len(tbl_df), len(tbl_df.columns), round(size_mb, 3)])
    for col_idx in range(1, 5):
        ws_res.column_dimensions[get_column_letter(col_idx)].width = 25

    wb.save(out_path)
    print(f"  Diccionario guardado: {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# 10. RESUMEN ESTADÍSTICO
# ─────────────────────────────────────────────────────────────────────────────

def print_summary(
    df_raw_len: int,
    df_after_len: int,
    fact: pd.DataFrame,
    df: pd.DataFrame,
    dim_ubicacion: pd.DataFrame,
    dim_tipo: pd.DataFrame,
    clima_ok: bool,
) -> None:
    sep = "─" * 60
    print(f"\n{sep}")
    print("  RESUMEN DEL PIPELINE")
    print(sep)
    print(f"  Filas en CSV crudo      : {df_raw_len:>10,}")
    print(f"  Filas tras filtros      : {df_after_len:>10,}")
    print(f"  Filas en fact_incidentes: {len(fact):>10,}")

    t_min = df["datetime_creacion"].min()
    t_max = df["datetime_creacion"].max()
    print(f"  Rango temporal          : {t_min.date()} → {t_max.date()}")

    print(f"\n  Top 10 alcaldías por volumen:")
    alcaldia_col = df["alcaldia_catalogo"].where(
        df["alcaldia_catalogo"] != "DESCONOCIDA", df["alcaldia_inicio"]
    )
    top_alc = alcaldia_col.value_counts().head(10)
    for alc, cnt in top_alc.items():
        print(f"    {alc:<35} {cnt:>8,}")

    print(f"\n  Top 5 tipos de incidente:")
    top_tipo = df["tipo_incidente_c4"].value_counts().head(5)
    for tipo, cnt in top_tipo.items():
        print(f"    {str(tipo):<40} {cnt:>8,}")

    print(f"\n  Distribución tiempo_respuesta_min:")
    tr = fact["tiempo_respuesta_min"].dropna().astype(float)
    for q, lbl in [(0.10, "p10"), (0.25, "p25"), (0.50, "p50"), (0.75, "p75"), (0.90, "p90")]:
        print(f"    {lbl}: {tr.quantile(q):.0f} min")

    print(f"\n  Cuadrantes únicos generados: {dim_ubicacion['cuadrante_id'].nunique():,}")

    if not clima_ok:
        print(f"\n  {'─'*56}")
        print("  ⚠️  dim_clima generada SIN datos de API.")
        print("     Para enriquecer: re-ejecutar el script cuando")
        print("     haya conectividad a archive-api.open-meteo.com")
        print(f"  {'─'*56}")
    print(sep)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="ETL C5 CDMX → esquema estrella")
    parser.add_argument("--input",  required=True, help="Path al CSV crudo")
    parser.add_argument("--output", required=True, help="Directorio raíz de salida")
    parser.add_argument("--sample", type=int, default=None, help="Número de filas a procesar")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output)
    cache_dir = input_path.parent.parent / "cache"

    # Crear estructura de directorios
    (output_dir / "HECHOS").mkdir(parents=True, exist_ok=True)
    (output_dir / "DIMENSIONES").mkdir(parents=True, exist_ok=True)
    (output_dir / "DICCIONARIO").mkdir(parents=True, exist_ok=True)

    print(f"\n{'═'*60}")
    print("  ETL: C5 CDMX → Esquema Estrella")
    print(f"{'═'*60}")
    print(f"  Input : {input_path}")
    print(f"  Output: {output_dir}")
    if args.sample:
        print(f"  Sample: {args.sample:,} filas")

    # ── Leer CSV ───────────────────────────────────────────────────────────────
    print("\n[1/8] Leyendo CSV crudo...")
    df_raw = _read_raw_csv(input_path, args.sample)
    n_raw = len(df_raw)

    # ── Preprocesar ────────────────────────────────────────────────────────────
    print("\n[2/8] Preprocesando...")
    df = preprocess(df_raw)
    n_after = len(df)

    # ── Dimensiones ────────────────────────────────────────────────────────────
    print("\n[3/8] Construyendo dim_tiempo...")
    dim_tiempo = build_dim_tiempo(df)
    print(f"  {len(dim_tiempo):,} registros horarios únicos")

    print("\n[4/8] Construyendo dim_ubicacion...")
    dim_ubicacion = build_dim_ubicacion(df)
    print(f"  {len(dim_ubicacion):,} registros (cuadrante, alcaldía, colonia)")

    print("\n[5/8] Construyendo dim_clima (Open-Meteo)...")
    dim_clima, clima_ok = build_dim_clima(df, cache_dir)
    print(f"  {len(dim_clima):,} horas en dim_clima  (api_ok={clima_ok})")

    print("\n[6/8] Construyendo dim_calendario...")
    dim_calendario = build_dim_calendario(df)
    print(f"  {len(dim_calendario):,} días")

    print("\n[7/8] Construyendo dim_tipo_incidente...")
    dim_tipo = build_dim_tipo_incidente(df)
    print(f"  {len(dim_tipo):,} combinaciones tipo/incidente/canal")

    # ── Fact ───────────────────────────────────────────────────────────────────
    print("\n[8/8] Construyendo fact_incidentes + integridad referencial...")
    fact = build_fact(df, dim_tiempo, dim_ubicacion, dim_clima, dim_calendario, dim_tipo)

    dims_dict = {
        "dim_tiempo": dim_tiempo,
        "dim_ubicacion": dim_ubicacion,
        "dim_clima": dim_clima,
        "dim_calendario": dim_calendario,
        "dim_tipo_incidente": dim_tipo,
    }
    fact, dims_dict = fix_referential_integrity(fact, dims_dict)

    # ── Guardar CSVs ───────────────────────────────────────────────────────────
    print("\n  Guardando archivos CSV...")
    fact.to_csv(output_dir / "HECHOS" / "fact_incidentes.csv", index=False)
    dims_dict["dim_tiempo"].to_csv(output_dir / "DIMENSIONES" / "dim_tiempo.csv", index=False)
    dims_dict["dim_ubicacion"].to_csv(output_dir / "DIMENSIONES" / "dim_ubicacion.csv", index=False)
    dims_dict["dim_clima"].to_csv(output_dir / "DIMENSIONES" / "dim_clima.csv", index=False)
    dims_dict["dim_calendario"].to_csv(output_dir / "DIMENSIONES" / "dim_calendario.csv", index=False)
    dims_dict["dim_tipo_incidente"].to_csv(output_dir / "DIMENSIONES" / "dim_tipo_incidente.csv", index=False)

    for name, tbl in [("fact_incidentes", fact)] + list(dims_dict.items()):
        path = output_dir / ("HECHOS" if name == "fact_incidentes" else "DIMENSIONES") / f"{name}.csv"
        size = path.stat().st_size / (1024 ** 2)
        print(f"    {name:<25} {len(tbl):>8,} filas  {size:>6.1f} MB")

    # ── Diccionario ─────────────────────────────────────────────────────────────
    print("\n  Construyendo diccionario_datos.xlsx...")
    build_diccionario(
        output_dir / "DICCIONARIO" / "diccionario_datos.xlsx",
        dims_dict,
        fact,
    )

    # ── Resumen ─────────────────────────────────────────────────────────────────
    print_summary(n_raw, n_after, fact, df, dims_dict["dim_ubicacion"], dim_tipo, clima_ok)


if __name__ == "__main__":
    main()
